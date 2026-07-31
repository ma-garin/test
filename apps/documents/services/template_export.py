"""成果物を Excel ひな型へ書き出す（要件 #62）。

旧実装の「安全モード」を引き継ぐ。ひな型は利用者が業務で使っている実ファイルで、
数式・書式・印刷設定が入っている。**マッピングで指定されたセル以外は触らない。**

書けなかったものは黙って落とさず、必ず `warnings` に理由を残す。
「出力できた」と「全部書けた」は別のことなので、画面でも区別して見せる。
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import date

from django.core.files.base import ContentFile
from django.utils import timezone

from apps.documents.models import Template, TemplateOutput

#: 数式が入っているセルは上書きしない。壊すと利用者側の集計が静かに狂う。
FORMULA_PREFIX = "="


@dataclass(frozen=True)
class ExportResult:
    """1 回の書き出し結果。"""

    output: TemplateOutput | None = None
    written: tuple[str, ...] = ()
    warnings: tuple[str, ...] = ()
    errors: tuple[str, ...] = ()

    @property
    def ok(self) -> bool:
        return self.output is not None

    @property
    def written_count(self) -> int:
        return len(self.written)


@dataclass
class _Values:
    """マッピングの項目名 → 実際に書き込む値。

    値は成果物と案件からしか作らない。ここで計算した数字を持ち込まないのは、
    本文と Excel で数字が食い違うのを防ぐため。
    """

    items: dict[str, object] = field(default_factory=dict)

    def get(self, name: str):
        return self.items.get(_normalize(name))


def _normalize(name: str) -> str:
    """項目名の表記揺れを吸収する。全角空白と前後空白だけ落とす。"""

    return str(name or "").replace("　", " ").strip()


def build_values(deliverable) -> _Values:
    """成果物 1 件から、書き込める値の辞書を作る。"""

    from apps.pmo.models import Approval

    project = deliverable.project
    # 「承認済み」の判断だけを見る。差し戻しの実施者を承認者として書くと誤解を生む。
    approval = (
        deliverable.approvals.filter(decision=Approval.Decision.APPROVED)
        .order_by("-created_at")
        .first()
        if deliverable.pk
        else None
    )
    rate = deliverable.correction_rate

    values = {
        "案件名": project.name,
        "案件コード": project.code,
        "案件ステータス": project.get_status_display(),
        "進捗率": project.progress_percent,
        "プロジェクトマネージャー": project.project_manager,
        "PMO担当": project.pmo_manager,
        "タイトル": deliverable.title,
        "種別": deliverable.get_kind_display(),
        "版": deliverable.version,
        "状態": deliverable.get_status_display(),
        "本文": deliverable.body or deliverable.ai_generated_body,
        "AI生成本文": deliverable.ai_generated_body,
        "作成日": timezone.localdate(deliverable.created_at)
        if deliverable.created_at
        else timezone.localdate(),
        "作成者": str(deliverable.created_by) if deliverable.created_by else "",
        "出力日": timezone.localdate(),
        "赤字率": rate if rate is not None else "",
        "承認者": str(approval.actor) if approval is not None and approval.actor else "",
        "承認日": timezone.localdate(approval.created_at) if approval is not None else "",
    }

    return _Values({_normalize(key): value for key, value in values.items()})


def _split_cell(spec: str) -> tuple[str | None, str]:
    """`Sheet1!B3` を（シート名, セル）へ分ける。シート名が無ければ None。"""

    text = str(spec or "").strip()

    if "!" in text:
        sheet, _, cell = text.partition("!")

        return sheet.strip().strip("'"), cell.strip()

    return None, text


def _target_sheet(workbook, sheet_name: str | None):
    if sheet_name is None:
        return workbook.active

    if sheet_name in workbook.sheetnames:
        return workbook[sheet_name]

    return None


def _writable_value(value):
    """openpyxl が受け付ける型へ寄せる。日付はそのまま、それ以外は文字列化する。"""

    if value is None:
        return ""

    if isinstance(value, (int, float, date)):
        return value

    return str(value)


def export(template: Template, deliverable, *, user=None) -> ExportResult:
    """ひな型へ成果物を書き出し、`TemplateOutput` として保存する。

    書き込み対象は `Template.field_mapping` に指定されたセルだけ。
    数式セル・存在しないシートは書かずに警告へ回す。
    """

    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover - 依存が入っていない環境向け
        return ExportResult(
            errors=(
                "openpyxl が導入されていません。"
                "`pip install -r requirements/ingest.txt` を実行してください。",
            )
        )

    mapping = template.field_mapping if isinstance(template.field_mapping, dict) else {}

    if not mapping:
        return ExportResult(errors=("項目マッピングが未設定です。書き込み先セルが決まりません。",))

    warnings: list[str] = []

    if template.mapping_status != Template.MappingStatus.APPROVED:
        # 未承認のマッピングでも出力は止めない。止めると承認前の確認ができない。
        warnings.append(
            f"項目マッピングが「{template.get_mapping_status_display()}」です。"
            "内容を確認してから配布してください。"
        )

    try:
        workbook = load_workbook(template.file.path)
    except FileNotFoundError:
        return ExportResult(errors=("ひな型ファイルが見つかりません。再登録してください。",))
    except Exception as exc:  # openpyxl は形式ごとに別の例外を投げる
        return ExportResult(errors=(f"ひな型を読み込めません: {exc}",))

    values = build_values(deliverable)
    written: list[str] = []

    for field_name, cell_spec in sorted(mapping.items()):
        value = values.get(field_name)

        if value is None:
            warnings.append(f"「{field_name}」に対応する値がありません。セルは空のままです。")
            continue

        sheet_name, cell_ref = _split_cell(cell_spec)
        sheet = _target_sheet(workbook, sheet_name)

        if sheet is None:
            warnings.append(f"「{field_name}」の書き込み先シート「{sheet_name}」がありません。")
            continue

        try:
            cell = sheet[cell_ref]
        except (ValueError, KeyError, TypeError):
            warnings.append(f"「{field_name}」のセル指定「{cell_spec}」を解釈できません。")
            continue

        if isinstance(cell.value, str) and cell.value.startswith(FORMULA_PREFIX):
            # 数式を消すと利用者側の集計が壊れる。上書きしない。
            warnings.append(f"「{field_name}」の書き込み先 {cell_spec} は数式です。上書きしません。")
            continue

        cell.value = _writable_value(value)
        written.append(f"{field_name} → {cell_spec}")

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    stamp = timezone.localtime().strftime("%Y%m%d-%H%M%S")
    filename = f"{template.name}_{deliverable.title}_{stamp}.xlsx".replace("/", "_")

    output = TemplateOutput.objects.create(
        template=template,
        project=deliverable.project,
        generated_by=user,
        warnings=warnings,
    )
    output.file.save(filename, ContentFile(buffer.read()), save=True)

    return ExportResult(output=output, written=tuple(written), warnings=tuple(warnings))
