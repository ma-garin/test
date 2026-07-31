"""Excel ひな型への成果物出力（要件 #62）。

実ファイルを読み書きするため MEDIA_ROOT を一時ディレクトリへ差し替える。
ひな型は openpyxl でその場で作る。バイナリをリポジトリへ置かない。
"""

from __future__ import annotations

import io
import tempfile

from django.core.files.base import ContentFile
from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import Workbook, load_workbook

from apps.accounts.constants import Role
from apps.accounts.models import Tenant, User
from apps.documents.models import Template, TemplateOutput
from apps.documents.services import template_export
from apps.pmo.models import Approval, Deliverable
from apps.projects.models import Project, ProjectMember

MEDIA_ROOT = tempfile.mkdtemp(prefix="verirag-test-export-")


def _workbook_bytes(*, with_formula: bool = False) -> bytes:
    """試験用のひな型。既存の値と数式を持たせ、壊れないことを確かめる。"""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "サマリ"
    sheet["A1"] = "案件名"
    sheet["A2"] = "進捗率"

    if with_formula:
        sheet["B5"] = "=SUM(B1:B4)"

    workbook.create_sheet("課題一覧")
    buffer = io.BytesIO()
    workbook.save(buffer)

    return buffer.getvalue()


@override_settings(MEDIA_ROOT=MEDIA_ROOT)
class TemplateExportTests(TestCase):
    def setUp(self) -> None:
        self.tenant = Tenant.objects.create(code="acme", name="ACME")
        self.user = User.objects.create_user(
            username="pmo",
            email="pmo@example.com",
            password="test-password",
            tenant=self.tenant,
            role=Role.PMO,
            display_name="PMO 太郎",
        )
        self.project = Project.objects.create(
            tenant=self.tenant, code="p1", name="基幹刷新", progress_percent=62
        )
        ProjectMember.objects.create(project=self.project, user=self.user, role_label="PMO")
        self.deliverable = Deliverable.objects.create(
            project=self.project,
            kind=Deliverable.Kind.WEEKLY_REPORT,
            title="第3週 週次報告",
            ai_generated_body="AIが書いた本文",
            body="人が直した本文",
            created_by=self.user,
        )

    def _template(self, *, mapping: dict, with_formula: bool = False, status=None) -> Template:
        template = Template.objects.create(
            tenant=self.tenant,
            name="週次報告ひな型",
            field_mapping=mapping,
            mapping_status=status or Template.MappingStatus.APPROVED,
        )
        template.file.save(
            "weekly.xlsx", ContentFile(_workbook_bytes(with_formula=with_formula)), save=True
        )

        return template

    def test_マッピングされたセルへ値を書き出す(self) -> None:
        template = self._template(mapping={"案件名": "B1", "進捗率": "B2", "本文": "課題一覧!A1"})

        result = template_export.export(template, self.deliverable, user=self.user)

        self.assertTrue(result.ok)
        self.assertEqual(result.written_count, 3)

        workbook = load_workbook(result.output.file.path)
        self.assertEqual(workbook["サマリ"]["B1"].value, "基幹刷新")
        self.assertEqual(workbook["サマリ"]["B2"].value, 62)
        # 確定本文があるならそちらを出す。AI生成本文をそのまま配布させない。
        self.assertEqual(workbook["課題一覧"]["A1"].value, "人が直した本文")

    def test_ひな型の既存の値を壊さない(self) -> None:
        template = self._template(mapping={"案件名": "B1"})

        result = template_export.export(template, self.deliverable)

        workbook = load_workbook(result.output.file.path)
        self.assertEqual(workbook["サマリ"]["A1"].value, "案件名")
        self.assertEqual(workbook["サマリ"]["A2"].value, "進捗率")

    def test_数式セルは上書きせず警告する(self) -> None:
        template = self._template(mapping={"案件名": "B5"}, with_formula=True)

        result = template_export.export(template, self.deliverable)

        self.assertTrue(result.ok)
        self.assertEqual(result.written_count, 0)
        self.assertTrue(any("数式" in w for w in result.warnings))

        workbook = load_workbook(result.output.file.path)
        self.assertEqual(workbook["サマリ"]["B5"].value, "=SUM(B1:B4)")

    def test_対応する値が無い項目は警告に残す(self) -> None:
        template = self._template(mapping={"存在しない項目": "B1"})

        result = template_export.export(template, self.deliverable)

        self.assertTrue(result.ok)
        self.assertEqual(result.written_count, 0)
        self.assertTrue(any("存在しない項目" in w for w in result.warnings))

    def test_存在しないシートを指定しても落ちない(self) -> None:
        template = self._template(mapping={"案件名": "無いシート!B1"})

        result = template_export.export(template, self.deliverable)

        self.assertTrue(result.ok)
        self.assertTrue(any("無いシート" in w for w in result.warnings))

    def test_マッピング未設定なら出力せず理由を返す(self) -> None:
        template = self._template(mapping={})

        result = template_export.export(template, self.deliverable)

        self.assertFalse(result.ok)
        self.assertTrue(any("マッピング" in e for e in result.errors))

    def test_未承認マッピングは出力するが警告する(self) -> None:
        template = self._template(
            mapping={"案件名": "B1"}, status=Template.MappingStatus.DRAFT
        )

        result = template_export.export(template, self.deliverable)

        self.assertTrue(result.ok)
        self.assertTrue(any("下書き" in w for w in result.warnings))

    def test_承認済みなら承認者と承認日を書ける(self) -> None:
        Approval.objects.create(
            deliverable=self.deliverable, actor=self.user, decision=Approval.Decision.APPROVED
        )
        template = self._template(mapping={"承認者": "B1"})

        result = template_export.export(template, self.deliverable)

        workbook = load_workbook(result.output.file.path)
        self.assertEqual(workbook["サマリ"]["B1"].value, "PMO 太郎")

    def test_差し戻しの実施者を承認者として書かない(self) -> None:
        Approval.objects.create(
            deliverable=self.deliverable, actor=self.user, decision=Approval.Decision.REJECTED
        )
        template = self._template(mapping={"承認者": "B1"})

        result = template_export.export(template, self.deliverable)

        workbook = load_workbook(result.output.file.path)
        # 空文字を書いたセルは openpyxl 上 None として読み出される。
        self.assertFalse(workbook["サマリ"]["B1"].value)


@override_settings(MEDIA_ROOT=MEDIA_ROOT)
class TemplateExportViewTests(TestCase):
    def setUp(self) -> None:
        self.tenant = Tenant.objects.create(code="acme", name="ACME")
        self.other_tenant = Tenant.objects.create(code="beta", name="BETA")
        self.user = User.objects.create_user(
            username="pmo",
            email="pmo@example.com",
            password="test-password",
            tenant=self.tenant,
            role=Role.PMO,
        )
        self.client.force_login(self.user)
        session = self.client.session
        session["tenant_id"] = str(self.tenant.pk)
        session.save()

        self.project = Project.objects.create(tenant=self.tenant, code="p1", name="基幹刷新")
        ProjectMember.objects.create(project=self.project, user=self.user, role_label="PMO")
        self.deliverable = Deliverable.objects.create(
            project=self.project, title="第3週 週次報告", body="本文"
        )
        self.template = Template.objects.create(
            tenant=self.tenant,
            name="週次報告ひな型",
            field_mapping={"案件名": "B1"},
            mapping_status=Template.MappingStatus.APPROVED,
        )
        self.template.file.save("weekly.xlsx", ContentFile(_workbook_bytes()), save=True)

    def test_画面から出力できる(self) -> None:
        response = self.client.post(
            reverse("pmo:deliverables"),
            {
                "action": "export",
                "deliverable": str(self.deliverable.pk),
                "template": str(self.template.pk),
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(TemplateOutput.objects.count(), 1)

    def test_他テナントのひな型は使えない(self) -> None:
        foreign = Template.objects.create(
            tenant=self.other_tenant, name="他社ひな型", field_mapping={"案件名": "B1"}
        )
        foreign.file.save("other.xlsx", ContentFile(_workbook_bytes()), save=True)

        response = self.client.post(
            reverse("pmo:deliverables"),
            {
                "action": "export",
                "deliverable": str(self.deliverable.pk),
                "template": str(foreign.pk),
            },
        )

        self.assertEqual(response.status_code, 404)
        self.assertEqual(TemplateOutput.objects.count(), 0)
